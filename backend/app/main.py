from datetime import date, datetime
from io import BytesIO
import json
import re
import time
import unicodedata
from fastapi import Depends, FastAPI, File, Form, HTTPException, Response, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook, load_workbook
from sqlalchemy import inspect, text
from sqlalchemy.orm import Session
from sqlalchemy.exc import OperationalError

from .auth import create_access_token, get_current_user, hash_password, require_admin, verify_password
from .config import get_settings
from .crud import (
    create_announcement,
    create_course,
    create_partner,
    create_user,
    delete_announcement,
    delete_course,
    delete_partner,
    get_portal_link,
    get_user_by_email,
    list_announcements,
    list_courses,
    list_partners,
    list_portal_links,
    list_users,
    update_announcement,
    update_course,
    update_partner,
    update_user,
)
from .database import Base, SessionLocal, engine, get_db
from . import models
from .models import Announcement, Course
from .schemas import (
    AnnouncementCreate,
    AnnouncementPublic,
    AnnouncementUpdate,
    CourseCreate,
    CoursePublic,
    CourseUpdate,
    LoginRequest,
    BulkUserResult,
    PartnerCreate,
    PartnerPublic,
    PartnerUpdate,
    PortalLinkPublic,
    UserImportApplyResult,
    UserImportPreview,
    UserImportRow,
    UserCreate,
    UserPasswordUpdate,
    UserPublic,
    UserUpdate,
)
from .seed import seed_all

settings = get_settings()

app = FastAPI(title="Sou Arte em Cuidados API")

origins = [origin.strip() for origin in settings.cors_origins.split(",") if origin.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins or ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup() -> None:
    for _ in range(10):
        try:
            Base.metadata.create_all(bind=engine)
            with engine.begin() as connection:
                inspector = inspect(connection)
                announcement_columns = {column["name"] for column in inspector.get_columns("announcements")}
                if "expires_at" not in announcement_columns:
                    connection.execute(text("ALTER TABLE announcements ADD COLUMN expires_at DATETIME NULL"))
                if "target_cities" not in announcement_columns:
                    connection.execute(text("ALTER TABLE announcements ADD COLUMN target_cities TEXT NULL"))
                if "target_professions" not in announcement_columns:
                    connection.execute(text("ALTER TABLE announcements ADD COLUMN target_professions TEXT NULL"))

                course_columns = {column["name"] for column in inspector.get_columns("courses")}
                if "target_cities" not in course_columns:
                    connection.execute(text("ALTER TABLE courses ADD COLUMN target_cities TEXT NULL"))
                if "target_professions" not in course_columns:
                    connection.execute(text("ALTER TABLE courses ADD COLUMN target_professions TEXT NULL"))

                partner_columns = {column["name"] for column in inspector.get_columns("partners")}
                if "target_cities" not in partner_columns:
                    connection.execute(text("ALTER TABLE partners ADD COLUMN target_cities TEXT NULL"))
                if "target_professions" not in partner_columns:
                    connection.execute(text("ALTER TABLE partners ADD COLUMN target_professions TEXT NULL"))

                user_columns = {column["name"] for column in inspector.get_columns("users")}
                if "city" not in user_columns:
                    connection.execute(text("ALTER TABLE users ADD COLUMN city VARCHAR(120) NULL"))
                if "uf" not in user_columns:
                    connection.execute(text("ALTER TABLE users ADD COLUMN uf VARCHAR(2) NULL"))
                if "admission_date" not in user_columns:
                    connection.execute(text("ALTER TABLE users ADD COLUMN admission_date DATE NULL"))
                if "profession" not in user_columns:
                    connection.execute(text("ALTER TABLE users ADD COLUMN profession VARCHAR(120) NULL"))
            with SessionLocal() as db:
                seed_all(db)
            return
        except OperationalError:
            time.sleep(2)
    Base.metadata.create_all(bind=engine)
    with SessionLocal() as db:
        seed_all(db)


CREATE_HEADERS = {
    "name": ["nome", "name"],
    "email": ["email", "e mail"],
    "password": ["senha", "password"],
    "role": ["perfil", "role", "tipo", "papel"],
}

DELETE_HEADERS = {
    "email": ["email", "e mail"],
}

SYNC_HEADERS = {
    "name": ["nome", "name"],
    "email": ["email", "e mail"],
    "city": ["cidade", "city"],
    "uf": ["uf", "estado"],
    "admission_date": ["dt admissao", "data admissao", "data de admissao", "admissao", "dt de admissao"],
    "profession": ["profissao", "cargo", "especialidade"],
}

SYNC_REQUIRED_FIELDS = ("name", "email")

MISSING_FIELD_LABELS = {
    "city": "cidade",
    "uf": "UF",
    "admission_date": "data de admissão",
    "profession": "profissão",
}

ROLE_ALIASES = {
    "admin": "admin",
    "administrador": "admin",
    "adm": "admin",
    "socio": "socio",
    "sócio": "socio",
    "associado": "socio",
}


def normalize_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value: object) -> str:
    text_value = normalize_value(value).lower()
    text_value = unicodedata.normalize("NFKD", text_value)
    text_value = "".join(character for character in text_value if not unicodedata.combining(character))
    text_value = re.sub(r"[^a-z0-9]+", " ", text_value).strip()
    return text_value


def normalize_optional_text(value: object) -> str | None:
    text_value = normalize_value(value)
    return text_value if text_value else None


def parse_optional_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text_value = normalize_value(value)
    if not text_value:
        return None

    for date_format in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text_value, date_format).date()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text_value).date()
    except ValueError:
        return None


def find_header_row_and_index(
    ws,
    header_aliases: dict[str, list[str]],
    required_fields: tuple[str, ...],
    max_scan_rows: int = 25,
) -> tuple[int, dict[str, int]]:
    normalized_aliases = {
        field: {normalize_header(alias) for alias in aliases}
        for field, aliases in header_aliases.items()
    }

    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        index_map: dict[str, int] = {}
        for column_index, cell_value in enumerate(row):
            header = normalize_header(cell_value)
            if not header:
                continue
            for field, aliases in normalized_aliases.items():
                if header in aliases and field not in index_map:
                    index_map[field] = column_index

        if all(field in index_map for field in required_fields):
            return row_index, index_map

    missing_label = ", ".join(required_fields)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=f"Arquivo inválido. Colunas obrigatórias ausentes: {missing_label}.",
    )


def get_header_index(ws, required_headers: dict[str, list[str]]) -> dict[str, int]:
    _, index_map = find_header_row_and_index(ws, required_headers, tuple(required_headers.keys()), max_scan_rows=5)
    return index_map


def parse_sync_rows(file: UploadFile) -> tuple[list[dict[str, object]], list[dict[str, object]], int]:
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Envie um arquivo .xlsx.")

    file.file.seek(0)
    workbook = load_workbook(file.file, data_only=True)
    sheet = workbook.active

    header_row, header_index = find_header_row_and_index(sheet, SYNC_HEADERS, SYNC_REQUIRED_FIELDS)

    processed = 0
    errors: list[dict[str, object]] = []
    seen_emails: set[str] = set()
    rows: list[dict[str, object]] = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        row_values = {
            field: row[index] if index < len(row) else None
            for field, index in header_index.items()
        }
        if all(normalize_value(value) == "" for value in row_values.values()):
            continue

        processed += 1

        name = normalize_value(row_values.get("name"))
        email = normalize_value(row_values.get("email")).lower()
        city = normalize_optional_text(row_values.get("city"))
        uf_value = normalize_optional_text(row_values.get("uf"))
        uf = uf_value.upper()[:2] if uf_value else None
        admission_date = parse_optional_date(row_values.get("admission_date"))
        profession = normalize_optional_text(row_values.get("profession"))

        if not name or not email:
            errors.append({"row": row_number, "message": "Nome e e-mail são obrigatórios."})
            continue

        if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            errors.append({"row": row_number, "message": "E-mail inválido."})
            continue

        if email in seen_emails:
            errors.append({"row": row_number, "message": "E-mail duplicado no arquivo."})
            continue

        seen_emails.add(email)
        rows.append(
            {
                "row": row_number,
                "name": name,
                "email": email,
                "city": city,
                "uf": uf,
                "admission_date": admission_date,
                "profession": profession,
            }
        )

    return rows, errors, processed


def missing_fields_for_completion(user: models.User, imported_row: dict[str, object]) -> list[str]:
    missing_fields: list[str] = []
    for field_name in ("city", "uf", "admission_date", "profession"):
        imported_value = imported_row.get(field_name)
        current_value = getattr(user, field_name)
        if (current_value is None or str(current_value).strip() == "") and imported_value is not None:
            missing_fields.append(MISSING_FIELD_LABELS[field_name])
    return missing_fields


def map_import_row(
    imported_row: dict[str, object],
    exists: bool,
    missing_fields: list[str],
) -> UserImportRow:
    return UserImportRow(
        row=int(imported_row["row"]),
        name=str(imported_row["name"]),
        email=str(imported_row["email"]),
        city=imported_row.get("city"),
        uf=imported_row.get("uf"),
        admission_date=imported_row.get("admission_date"),
        profession=imported_row.get("profession"),
        exists=exists,
        needs_completion=len(missing_fields) > 0,
        missing_fields=missing_fields,
    )


def delete_user_with_relations(db: Session, user: models.User) -> None:
    db.query(Announcement).filter(Announcement.created_by == user.id).update(
        {Announcement.created_by: None},
        synchronize_session=False,
    )
    db.query(Course).filter(Course.created_by == user.id).update(
        {Course.created_by: None},
        synchronize_session=False,
    )
    db.delete(user)


@app.post("/auth/login", response_model=UserPublic)
def login(payload: LoginRequest, response: Response, db: Session = Depends(get_db)):
    user = get_user_by_email(db, payload.email)
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciais inválidas.")
    if not user.active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Usuário inativo.")

    token = create_access_token(user.id)
    response.set_cookie(
        key=settings.cookie_name,
        value=token,
        httponly=True,
        secure=settings.cookie_secure,
        samesite="lax",
        max_age=settings.access_token_expire_minutes * 60,
    )
    user.last_login_at = datetime.utcnow()
    db.commit()
    return user


@app.post("/auth/logout")
def logout(response: Response):
    response.delete_cookie(settings.cookie_name)
    return {"message": "Logout efetuado."}


@app.get("/auth/me", response_model=UserPublic)
def me(current_user=Depends(get_current_user)):
    return current_user


@app.get("/admin/users", response_model=list[UserPublic])
def admin_list_users(_: str = Depends(require_admin), db: Session = Depends(get_db)):
    return list_users(db)


@app.post("/admin/users", response_model=UserPublic)
def admin_create_user(payload: UserCreate, _: str = Depends(require_admin), db: Session = Depends(get_db)):
    if get_user_by_email(db, payload.email):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="E-mail já cadastrado.")
    return create_user(
        db,
        payload.name,
        payload.email,
        payload.password,
        payload.role,
        payload.city,
        payload.uf,
        payload.admission_date,
        payload.profession,
    )


@app.patch("/admin/users/{user_id}", response_model=UserPublic)
def admin_update_user(user_id: str, payload: UserUpdate, _: str = Depends(require_admin), db: Session = Depends(get_db)):
    existing = db.query(models.User).filter(models.User.id == user_id).first()
    if not existing:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuário não encontrado.")
    if payload.email and payload.email != existing.email and get_user_by_email(db, payload.email):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="E-mail já cadastrado.")
    return update_user(
        db,
        existing,
        payload.name,
        payload.email,
        payload.role,
        payload.active,
        payload.city,
        payload.uf,
        payload.admission_date,
        payload.profession,
    )


@app.patch("/admin/users/{user_id}/password", response_model=UserPublic)
def admin_update_user_password(
    user_id: str,
    payload: UserPasswordUpdate,
    _: str = Depends(require_admin),
    db: Session = Depends(get_db),
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuário não encontrado.")
    if not payload.password.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Senha inválida.")
    user.password_hash = hash_password(payload.password.strip())
    db.commit()
    db.refresh(user)
    return user


@app.delete("/admin/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def admin_delete_user(
    user_id: str,
    current_user=Depends(require_admin),
    db: Session = Depends(get_db),
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuário não encontrado.")
    if user.id == current_user.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Não é possível excluir o usuário logado.")
    try:
        delete_user_with_relations(db, user)
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Erro ao excluir usuário.") from exc
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/admin/users/templates/create")
def admin_users_template_create(_: str = Depends(require_admin)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "usuarios"
    sheet.append(["nome", "email", "senha", "perfil"])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo-usuarios-criacao.xlsx"'},
    )


@app.get("/admin/users/templates/delete")
def admin_users_template_delete(_: str = Depends(require_admin)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "usuarios"
    sheet.append(["email"])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo-usuarios-exclusao.xlsx"'},
    )


@app.post("/admin/users/bulk-create", response_model=BulkUserResult)
def admin_bulk_create_users(
    file: UploadFile = File(...),
    _: str = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Envie um arquivo .xlsx.")

    file.file.seek(0)
    workbook = load_workbook(file.file, data_only=True)
    sheet = workbook.active
    header_index = get_header_index(sheet, CREATE_HEADERS)

    processed = 0
    created = 0
    skipped = 0
    errors: list[dict[str, object]] = []
    seen_emails: set[str] = set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {field: row[index] if index < len(row) else None for field, index in header_index.items()}
        if all(normalize_value(value) == "" for value in values.values()):
            continue

        processed += 1
        name = normalize_value(values["name"])
        email = normalize_value(values["email"]).lower()
        password = normalize_value(values["password"])
        role_raw = normalize_value(values["role"]).lower()
        role = ROLE_ALIASES.get(role_raw)

        if not name or not email or not password or not role:
            errors.append({"row": row_number, "message": "Campos obrigatórios ausentes ou perfil inválido."})
            skipped += 1
            continue

        if email in seen_emails:
            errors.append({"row": row_number, "message": "E-mail duplicado no arquivo."})
            skipped += 1
            continue

        seen_emails.add(email)

        if get_user_by_email(db, email):
            errors.append({"row": row_number, "message": "E-mail já cadastrado."})
            skipped += 1
            continue

        try:
            create_user(db, name, email, password, role)
            created += 1
        except Exception:
            db.rollback()
            errors.append({"row": row_number, "message": "Erro ao criar usuário."})
            skipped += 1

    return BulkUserResult(processed=processed, created=created, skipped=skipped, errors=errors)


@app.post("/admin/users/bulk-delete", response_model=BulkUserResult)
def admin_bulk_delete_users(
    file: UploadFile = File(...),
    current_user=Depends(require_admin),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Envie um arquivo .xlsx.")

    file.file.seek(0)
    workbook = load_workbook(file.file, data_only=True)
    sheet = workbook.active
    header_index = get_header_index(sheet, DELETE_HEADERS)

    processed = 0
    deleted = 0
    skipped = 0
    errors: list[dict[str, object]] = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        email_value = row[header_index["email"]] if header_index["email"] < len(row) else None
        email = normalize_value(email_value).lower()
        if not email:
            continue

        processed += 1
        user = get_user_by_email(db, email)
        if not user:
            errors.append({"row": row_number, "message": "Usuário não encontrado."})
            skipped += 1
            continue

        if user.id == current_user.id:
            errors.append({"row": row_number, "message": "Não é possível excluir o usuário logado."})
            skipped += 1
            continue

        try:
            delete_user_with_relations(db, user)
            db.commit()
            deleted += 1
        except Exception:
            db.rollback()
            errors.append({"row": row_number, "message": "Erro ao excluir usuário."})
            skipped += 1

    return BulkUserResult(processed=processed, deleted=deleted, skipped=skipped, errors=errors)


@app.post("/admin/users/import/preview", response_model=UserImportPreview)
def admin_users_import_preview(
    file: UploadFile = File(...),
    current_user=Depends(require_admin),
    db: Session = Depends(get_db),
):
    rows, errors, processed = parse_sync_rows(file)
    users = list_users(db)
    users_by_email = {user.email.lower(): user for user in users}
    imported_emails = {str(row["email"]).lower() for row in rows}

    new_rows: list[UserImportRow] = []
    existing_rows: list[UserImportRow] = []
    completion_rows: list[UserImportRow] = []

    for row in rows:
        existing_user = users_by_email.get(str(row["email"]).lower())
        if existing_user:
            missing_fields = missing_fields_for_completion(existing_user, row)
            mapped_row = map_import_row(row, exists=True, missing_fields=missing_fields)
            existing_rows.append(mapped_row)
            if missing_fields:
                completion_rows.append(mapped_row)
            continue
        new_rows.append(map_import_row(row, exists=False, missing_fields=[]))

    missing_in_file = [
        user
        for user in users
        if user.role == "socio" and user.id != current_user.id and user.email.lower() not in imported_emails
    ]
    missing_in_file.sort(key=lambda user: user.name.lower())

    return UserImportPreview(
        processed=processed,
        valid_rows=len(rows),
        new_rows=new_rows,
        existing_rows=existing_rows,
        completion_rows=completion_rows,
        missing_in_file=missing_in_file,
        errors=errors,
    )


@app.post("/admin/users/import/apply", response_model=UserImportApplyResult)
def admin_users_import_apply(
    file: UploadFile = File(...),
    role_map_json: str = Form("{}"),
    create_new: bool = True,
    fill_missing_data: bool = True,
    delete_missing_users: bool = False,
    current_user=Depends(require_admin),
    db: Session = Depends(get_db),
):
    rows, parsing_errors, processed = parse_sync_rows(file)
    users = list_users(db)
    users_by_email = {user.email.lower(): user for user in users}
    imported_emails = {str(row["email"]).lower() for row in rows}

    created = 0
    updated = 0
    deleted = 0
    skipped = len(parsing_errors)
    errors = list(parsing_errors)

    role_map: dict[str, str] = {}
    try:
        parsed_role_map = json.loads(role_map_json or "{}")
        if isinstance(parsed_role_map, dict):
            for key, value in parsed_role_map.items():
                email_key = str(key).strip().lower()
                role_value = str(value).strip().lower()
                if not email_key:
                    continue
                role_map[email_key] = "admin" if role_value == "admin" else "socio"
    except json.JSONDecodeError:
        pass

    for row in rows:
        email = str(row["email"]).lower()
        existing_user = users_by_email.get(email)

        if existing_user is None:
            if not create_new:
                skipped += 1
                continue
            try:
                created_user = create_user(
                    db,
                    str(row["name"]),
                    email,
                    settings.import_user_default_password,
                    role_map.get(email, "socio"),
                    row.get("city"),
                    row.get("uf"),
                    row.get("admission_date"),
                    row.get("profession"),
                )
                users_by_email[email] = created_user
                created += 1
            except Exception:
                db.rollback()
                errors.append({"row": int(row["row"]), "message": "Erro ao criar usuário."})
                skipped += 1
            continue

        if not fill_missing_data:
            continue

        changed = False
        if not existing_user.city and row.get("city"):
            existing_user.city = row.get("city")
            changed = True
        if not existing_user.uf and row.get("uf"):
            existing_user.uf = row.get("uf")
            changed = True
        if not existing_user.admission_date and row.get("admission_date"):
            existing_user.admission_date = row.get("admission_date")
            changed = True
        if not existing_user.profession and row.get("profession"):
            existing_user.profession = row.get("profession")
            changed = True

        if changed:
            try:
                db.commit()
                db.refresh(existing_user)
                updated += 1
            except Exception:
                db.rollback()
                errors.append({"row": int(row["row"]), "message": "Erro ao complementar dados do usuário."})
                skipped += 1

    if delete_missing_users:
        removable_users = [
            user
            for user in users_by_email.values()
            if user.role == "socio" and user.id != current_user.id and user.email.lower() not in imported_emails
        ]
        for user in removable_users:
            try:
                delete_user_with_relations(db, user)
                db.commit()
                deleted += 1
            except Exception:
                db.rollback()
                errors.append({"row": 0, "message": f"Erro ao excluir usuário: {user.email}."})
                skipped += 1

    return UserImportApplyResult(
        processed=processed,
        created=created,
        updated=updated,
        deleted=deleted,
        skipped=skipped,
        errors=errors,
    )


@app.get("/admin/announcements", response_model=list[AnnouncementPublic])
def admin_list_announcements(_: str = Depends(require_admin), db: Session = Depends(get_db)):
    return list_announcements(db, only_active=False)


@app.post("/admin/announcements", response_model=AnnouncementPublic)
def admin_create_announcement(
    payload: AnnouncementCreate,
    current_user=Depends(require_admin),
    db: Session = Depends(get_db),
):
    return create_announcement(
        db,
        payload.title,
        payload.body,
        payload.published_at,
        payload.expires_at,
        payload.target_cities,
        payload.target_professions,
        current_user.id,
    )


@app.put("/admin/announcements/{announcement_id}", response_model=AnnouncementPublic)
def admin_update_announcement(
    announcement_id: str,
    payload: AnnouncementUpdate,
    _: str = Depends(require_admin),
    db: Session = Depends(get_db),
):
    announcement = db.query(Announcement).filter(Announcement.id == announcement_id).first()
    if not announcement:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comunicado não encontrado.")
    return update_announcement(
        db,
        announcement,
        payload.title,
        payload.body,
        payload.published_at,
        payload.expires_at,
        payload.target_cities,
        payload.target_professions,
    )


@app.delete("/admin/announcements/{announcement_id}", status_code=status.HTTP_204_NO_CONTENT)
def admin_delete_announcement(announcement_id: str, _: str = Depends(require_admin), db: Session = Depends(get_db)):
    announcement = db.query(Announcement).filter(Announcement.id == announcement_id).first()
    if not announcement:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Comunicado não encontrado.")
    delete_announcement(db, announcement)
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/admin/courses", response_model=list[CoursePublic])
def admin_list_courses(_: str = Depends(require_admin), db: Session = Depends(get_db)):
    return list_courses(db, only_active=False)


@app.post("/admin/courses", response_model=CoursePublic)
def admin_create_course(
    payload: CourseCreate,
    current_user=Depends(require_admin),
    db: Session = Depends(get_db),
):
    return create_course(
        db,
        payload.title,
        payload.description,
        payload.image_url,
        payload.access_url,
        payload.target_cities,
        payload.target_professions,
        current_user.id,
    )


@app.put("/admin/courses/{course_id}", response_model=CoursePublic)
def admin_update_course(
    course_id: str,
    payload: CourseUpdate,
    _: str = Depends(require_admin),
    db: Session = Depends(get_db),
):
    course = db.query(Course).filter(Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Curso não encontrado.")
    return update_course(
        db,
        course,
        payload.title,
        payload.description,
        payload.image_url,
        payload.access_url,
        payload.target_cities,
        payload.target_professions,
    )


@app.delete("/admin/courses/{course_id}", status_code=status.HTTP_204_NO_CONTENT)
def admin_delete_course(course_id: str, _: str = Depends(require_admin), db: Session = Depends(get_db)):
    course = db.query(Course).filter(Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Curso não encontrado.")
    delete_course(db, course)
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/admin/partners", response_model=list[PartnerPublic])
def admin_list_partners(_: str = Depends(require_admin), db: Session = Depends(get_db)):
    return list_partners(db, only_active=False)


@app.post("/admin/partners", response_model=PartnerPublic)
def admin_create_partner(payload: PartnerCreate, _: str = Depends(require_admin), db: Session = Depends(get_db)):
    return create_partner(
        db,
        payload.name,
        payload.description,
        payload.link_url,
        payload.logo_url,
        payload.target_cities,
        payload.target_professions,
    )


@app.put("/admin/partners/{partner_id}", response_model=PartnerPublic)
def admin_update_partner(
    partner_id: str,
    payload: PartnerUpdate,
    _: str = Depends(require_admin),
    db: Session = Depends(get_db),
):
    partner = db.query(models.Partner).filter(models.Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Parceiro não encontrado.")
    return update_partner(
        db,
        partner,
        payload.name,
        payload.description,
        payload.link_url,
        payload.logo_url,
        payload.target_cities,
        payload.target_professions,
    )


@app.delete("/admin/partners/{partner_id}", status_code=status.HTTP_204_NO_CONTENT)
def admin_delete_partner(partner_id: str, _: str = Depends(require_admin), db: Session = Depends(get_db)):
    partner = db.query(models.Partner).filter(models.Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Parceiro não encontrado.")
    delete_partner(db, partner)
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/portal/announcements", response_model=list[AnnouncementPublic])
def portal_announcements(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    return list_announcements(db, only_active=True, only_visible=True, user=current_user)


@app.get("/portal/courses", response_model=list[CoursePublic])
def portal_courses(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    return list_courses(db, only_active=True, user=current_user)


@app.get("/portal/links", response_model=list[PortalLinkPublic])
def portal_links(_: str = Depends(get_current_user), db: Session = Depends(get_db)):
    return list_portal_links(db)


@app.get("/portal/links/{slug}", response_model=PortalLinkPublic)
def portal_link(slug: str, _: str = Depends(get_current_user), db: Session = Depends(get_db)):
    link = get_portal_link(db, slug)
    if not link:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Link não encontrado.")
    return link


@app.get("/portal/partners", response_model=list[PartnerPublic])
def portal_partners(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    return list_partners(db, only_active=True, user=current_user)
